"""
Analytics & Reporting Service.

Provides:
- Dashboard summary statistics
- Time-series violation trends
- Camera/zone heatmaps
- Repeat offender tracking
- PDF / CSV / Excel report export
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy import func, select, and_, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.logging import get_logger
from app.models.violation import (
    ViolationRecord, DetectionSession, RepeatOffender,
    ViolationCategory, VehicleType, SeverityLevel,
)
from app.schemas.violation import (
    AnalyticsSummary, ViolationCategoryStat, HourlyDistribution, DailyTrend,
)

logger = get_logger(__name__)


class AnalyticsService:
    """Reads violation data and produces analytical summaries and reports."""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ─── Dashboard Summary ─────────────────────────────────────────────────────

    async def get_summary(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        camera_id: Optional[str] = None,
    ) -> AnalyticsSummary:
        """Generate comprehensive analytics summary for the dashboard."""
        if end_date is None:
            end_date = datetime.utcnow()
        if start_date is None:
            start_date = end_date - timedelta(days=30)

        filters = [
            ViolationRecord.created_at >= start_date,
            ViolationRecord.created_at <= end_date,
        ]
        if camera_id:
            filters.append(ViolationRecord.camera_id == camera_id)

        where = and_(*filters)

        # Total violations
        total_q = await self.db.execute(
            select(func.count()).select_from(ViolationRecord).where(where)
        )
        total_violations = total_q.scalar_one() or 0

        # Total sessions
        session_filters = [
            DetectionSession.created_at >= start_date,
            DetectionSession.created_at <= end_date,
        ]
        if camera_id:
            session_filters.append(DetectionSession.camera_id == camera_id)

        sess_q = await self.db.execute(
            select(func.count()).select_from(DetectionSession).where(and_(*session_filters))
        )
        total_sessions = sess_q.scalar_one() or 0

        # Violations by category
        cat_q = await self.db.execute(
            select(ViolationRecord.violation_category, func.count().label("cnt"))
            .where(where)
            .group_by(ViolationRecord.violation_category)
        )
        cat_rows = cat_q.all()
        violations_by_category = [
            ViolationCategoryStat(
                category=row.violation_category,
                count=row.cnt,
                percentage=round(row.cnt / max(total_violations, 1) * 100, 1),
            )
            for row in cat_rows
        ]

        # Violations by vehicle type
        vtype_q = await self.db.execute(
            select(ViolationRecord.vehicle_type, func.count().label("cnt"))
            .where(where)
            .group_by(ViolationRecord.vehicle_type)
        )
        violations_by_vehicle = {
            row.vehicle_type.value if row.vehicle_type else "unknown": row.cnt
            for row in vtype_q.all()
        }

        # Hourly distribution
        hour_q = await self.db.execute(
            select(
                func.extract("hour", ViolationRecord.created_at).label("hr"),
                func.count().label("cnt"),
            )
            .where(where)
            .group_by(text("hr"))
            .order_by(text("hr"))
        )
        hourly_distribution = [
            HourlyDistribution(hour=int(row.hr), count=row.cnt)
            for row in hour_q.all()
        ]

        # Daily trend
        daily_q = await self.db.execute(
            select(
                func.date_trunc("day", ViolationRecord.created_at).label("day"),
                func.count().label("cnt"),
            )
            .where(where)
            .group_by(text("day"))
            .order_by(text("day"))
        )
        daily_trend = [
            DailyTrend(
                date=row.day.strftime("%Y-%m-%d"),
                count=row.cnt,
            )
            for row in daily_q.all()
        ]

        # Avg confidence
        conf_q = await self.db.execute(
            select(func.avg(ViolationRecord.confidence)).where(where)
        )
        avg_conf = float(conf_q.scalar_one() or 0.0)

        # Top offending plates
        plate_q = await self.db.execute(
            select(ViolationRecord.plate_number, func.count().label("cnt"))
            .where(and_(where, ViolationRecord.plate_number.isnot(None)))
            .group_by(ViolationRecord.plate_number)
            .order_by(func.count().desc())
            .limit(10)
        )
        top_plates = [
            {"plate": row.plate_number, "count": row.cnt}
            for row in plate_q.all()
        ]

        # High-risk cameras
        cam_q = await self.db.execute(
            select(ViolationRecord.camera_id, func.count().label("cnt"))
            .where(and_(where, ViolationRecord.camera_id.isnot(None)))
            .group_by(ViolationRecord.camera_id)
            .order_by(func.count().desc())
            .limit(5)
        )
        high_risk_cameras = [
            {"camera_id": row.camera_id, "count": row.cnt}
            for row in cam_q.all()
        ]

        return AnalyticsSummary(
            period_start=start_date,
            period_end=end_date,
            total_violations=total_violations,
            total_sessions=total_sessions,
            total_vehicles_detected=total_sessions,  # approximation
            violations_by_category=violations_by_category,
            violations_by_vehicle_type=violations_by_vehicle,
            hourly_distribution=hourly_distribution,
            daily_trend=daily_trend,
            avg_confidence=avg_conf,
            top_offending_plates=top_plates,
            high_risk_cameras=high_risk_cameras,
        )

    # ─── Report Export ─────────────────────────────────────────────────────────

    async def export_csv(
        self,
        start_date: datetime,
        end_date: datetime,
        filters: Optional[Dict] = None,
    ) -> bytes:
        """Export violations as CSV bytes."""
        records = await self._fetch_records(start_date, end_date, filters)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "ID", "Session ID", "Violation Category", "Vehicle Type",
            "Severity", "Confidence", "Plate Number", "Camera ID",
            "Location", "Challan Issued", "Timestamp",
        ])

        for r in records:
            writer.writerow([
                str(r.id),
                str(r.session_id),
                r.violation_category.value,
                r.vehicle_type.value if r.vehicle_type else "",
                r.severity.value,
                f"{r.confidence:.4f}",
                r.plate_number or "",
                r.camera_id or "",
                r.location_label or "",
                r.challan_issued,
                r.created_at.isoformat(),
            ])

        return output.getvalue().encode("utf-8")

    async def export_excel(
        self,
        start_date: datetime,
        end_date: datetime,
        filters: Optional[Dict] = None,
    ) -> bytes:
        """Export violations as Excel bytes using openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed")
            raise

        records = await self._fetch_records(start_date, end_date, filters)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Violations"

        # Header styling
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        headers = [
            "ID", "Session ID", "Category", "Vehicle Type",
            "Severity", "Confidence", "Plate", "Camera", "Location",
            "Challan", "Timestamp",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, r in enumerate(records, 2):
            ws.append([
                str(r.id),
                str(r.session_id),
                r.violation_category.value,
                r.vehicle_type.value if r.vehicle_type else "",
                r.severity.value,
                round(r.confidence, 4),
                r.plate_number or "",
                r.camera_id or "",
                r.location_label or "",
                "Yes" if r.challan_issued else "No",
                r.created_at.isoformat(),
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def export_pdf(
        self,
        start_date: datetime,
        end_date: datetime,
        filters: Optional[Dict] = None,
    ) -> bytes:
        """Export summary report as PDF using reportlab."""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
        except ImportError:
            logger.error("reportlab not installed")
            raise

        summary = await self.get_summary(start_date, end_date)
        records = await self._fetch_records(start_date, end_date, filters)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        story.append(Paragraph("Traffic Violation Detection Report", styles["Title"]))
        story.append(Paragraph(
            f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 12))

        # Summary table
        summary_data = [
            ["Metric", "Value"],
            ["Total Violations", str(summary.total_violations)],
            ["Total Sessions", str(summary.total_sessions)],
            ["Average Confidence", f"{summary.avg_confidence:.1%}"],
        ]
        for cat in summary.violations_by_category:
            summary_data.append([
                cat.category.value.replace("_", " ").title(),
                f"{cat.count} ({cat.percentage}%)",
            ])

        t = Table(summary_data, colWidths=[250, 200])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F7")]),
        ]))
        story.append(t)
        story.append(Spacer(1, 24))

        # Records table
        story.append(Paragraph("Violation Records", styles["Heading2"]))
        rec_data = [["Category", "Vehicle", "Severity", "Plate", "Timestamp"]]
        for r in records[:200]:  # Cap at 200 rows for PDF
            rec_data.append([
                r.violation_category.value.replace("_", " ").title(),
                r.vehicle_type.value if r.vehicle_type else "N/A",
                r.severity.value.upper(),
                r.plate_number or "N/A",
                r.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        rt = Table(rec_data, colWidths=[130, 80, 70, 80, 110])
        rt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        story.append(rt)

        doc.build(story)
        return buffer.getvalue()

    # ─── Internal Helpers ──────────────────────────────────────────────────────

    async def _fetch_records(
        self,
        start_date: datetime,
        end_date: datetime,
        filters: Optional[Dict] = None,
    ) -> List[ViolationRecord]:
        """Fetch violation records with optional filters."""
        conditions = [
            ViolationRecord.created_at >= start_date,
            ViolationRecord.created_at <= end_date,
        ]
        if filters:
            if filters.get("violation_category"):
                conditions.append(ViolationRecord.violation_category == filters["violation_category"])
            if filters.get("vehicle_type"):
                conditions.append(ViolationRecord.vehicle_type == filters["vehicle_type"])
            if filters.get("plate_number"):
                conditions.append(ViolationRecord.plate_number.ilike(f"%{filters['plate_number']}%"))
            if filters.get("camera_id"):
                conditions.append(ViolationRecord.camera_id == filters["camera_id"])

        q = await self.db.execute(
            select(ViolationRecord)
            .where(and_(*conditions))
            .order_by(ViolationRecord.created_at.desc())
            .limit(10000)
        )
        return list(q.scalars().all())

    async def update_repeat_offender(self, plate_number: str, violation: Any):
        """Track and update repeat offender records."""
        q = await self.db.execute(
            select(RepeatOffender).where(RepeatOffender.plate_number == plate_number)
        )
        offender = q.scalar_one_or_none()

        if offender:
            offender.total_violations += 1
            offender.last_violation_at = datetime.utcnow()
            offender.last_violation_category = violation.violation_category
            offender.risk_score = min(1.0, offender.total_violations / 10.0)
            if offender.total_violations >= 3:
                offender.flagged = True
        else:
            offender = RepeatOffender(
                plate_number=plate_number,
                total_violations=1,
                last_violation_at=datetime.utcnow(),
                last_violation_category=violation.violation_category,
                risk_score=0.1,
            )
            self.db.add(offender)
